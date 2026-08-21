import os
import sys
import time
import io
import math

# 🚨 Windows 한글 인코딩 충돌 방지
os.environ["PYTHONIOENCODING"] = "utf-8"
os.environ["PYTHONUTF8"] = "1"
if sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
if sys.stderr.encoding.lower() != 'utf-8':
    sys.stderr.reconfigure(encoding='utf-8')

import streamlit as st
import pandas as pd
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import google.generativeai as genai

# ==========================================
# 🔐 API Key 연동
# ==========================================
try:
    API_KEY = st.secrets["GEMINI_API_KEY"]
except:
    API_KEY = None

# ==========================================
# 1. Core Engineering Logic
# ==========================================
class UtilityEngineeringEngine:
    def __init__(self):
        self.price_db = {
            "MCCB_50AF": 45000, "MCCB_125AF": 120000, "MCCB_250AF": 250000,
            "MCCB_400AF": 450000, "MCCB_630AF": 750000, "MCCB_800AF": 1100000, "MCCB_1000AF": 1500000,
            "Cable_Tray_W300": 25000
        }
        self.cable_data = [
            (2.5, 31, 18.5, 24, 8.91, 0.106), (4, 42, 25, 33, 5.57, 0.101),
            (6, 54, 32, 43, 3.71, 0.096), (10, 75, 44, 58, 2.24, 0.090),
            (16, 100, 59, 79, 1.41, 0.086), (25, 127, 77, 105, 0.889, 0.083),
            (35, 158, 96, 130, 0.641, 0.081), (50, 192, 117, 161, 0.473, 0.080),
            (70, 246, 149, 204, 0.328, 0.077), (95, 298, 180, 246, 0.236, 0.076),
            (120, 346, 208, 285, 0.187, 0.074), (150, 399, 236, 328, 0.153, 0.074),
            (185, 456, 268, 372, 0.122, 0.074), (240, 538, 315, 434, 0.093, 0.073),
            (300, 621, 363, 500, 0.074, 0.073)
        ]

    def step1_advanced_sizing(self, load_kw: float, distance_m: float, power_factor: float = 0.9, demand_factor: float = 1.0, is_continuous: bool = True, power_type: str = "3상4선", install_method: str = "E", temp_c: float = 30.0, is_critical_load: bool = False, tr_capacity_kva: float = 1000.0) -> dict:
        if "3상4선" in power_type: v_line, v_base, phase_type, b_coeff = 380, 220, 3, math.sqrt(3)
        elif "단상" in power_type: v_line, v_base, phase_type, b_coeff = 220, 220, 1, 2.0
        else: v_line, v_base, phase_type, b_coeff = 380, 380, 3, math.sqrt(3)

        # 1. 정격 전류 및 여유율 적용
        applied_kw = load_kw * demand_factor
        current = (applied_kw * 1000) / (math.sqrt(3) * v_line * power_factor) if phase_type == 3 else (applied_kw * 1000) / (v_line * power_factor)
        margin = 1.25 if is_continuous else 1.0
        target_cb = current * margin
        cb_standard = [30, 50, 100, 125, 200, 250, 400, 630, 800, 1000]
        selected_cb = next((cb for cb in cb_standard if cb >= target_cb), 1000)

        # 2. 고장전류(단락전류) 산출 및 차단기 차단용량(kA) 검증 (%Z = 5.0% 가정)
        tr_rated_current = tr_capacity_kva / (math.sqrt(3) * (v_line / 1000)) if tr_capacity_kva > 0 else 1000 / (math.sqrt(3) * 0.38)
        short_circuit_current_ka = (100 / 5.0) * tr_rated_current / 1000
        recommended_ka = 50 if short_circuit_current_ka > 25 else 25

        # 3. 무정전전원(UPS) 연계 여부 판단
        power_source = "UPS 또는 C-UPS 판넬 연계 필수" if is_critical_load else "일반 TR 판넬 연계"

        # 4. 온도 및 환경 계수 적용
        temp_factor = 1.0 if temp_c <= 30 else (0.96 if temp_c <= 35 else (0.91 if temp_c <= 40 else 0.82))
        method_idx = 2 if "A1" in install_method else (3 if "C" in install_method else 1)
        sin_theta = math.sqrt(1 - power_factor**2)

        # 5. KEC 정밀 케이블 Sizing
        optimal_sq, final_v_drop = None, 999.0
        for row in self.cable_data:
            sq, base_amp, r_val, x_val = row[0], row[method_idx], row[4], row[5]
            adjusted_amp = base_amp * temp_factor
            
            if adjusted_amp >= selected_cb:
                e_drop = (b_coeff * current * distance_m * (r_val * power_factor + x_val * sin_theta)) / 1000
                v_drop_p = (e_drop / v_base) * 100
                if v_drop_p <= 3.0:
                    optimal_sq = sq
                    final_v_drop = v_drop_p
                    break

        return {
            "load_current_A": round(current, 1),
            "selected_breaker_AT": f"{selected_cb}AT",
            "required_breaker_kA": f"최소 {recommended_ka}kA 이상 (단락전류 {round(short_circuit_current_ka, 1)}kA)",
            "recommended_power_source": power_source,
            "optimal_cable_SQ": optimal_sq if optimal_sq else "규격 초과 (다조 포설 필요)",
            "voltage_drop_percent": round(final_v_drop, 2)
        }

    def step2_evaluate_capacity(self, tr_capacity_kva: float, current_load_kw: float, add_power_kw: float) -> dict:
        expected_load = current_load_kw + add_power_kw
        expected_rate = (expected_load / tr_capacity_kva) * 100
        is_safe = expected_rate <= 80
        analysis = "적합" if is_safe else f"부하율 {expected_rate:.1f}%로 위험."
        return {"tr_capacity_kva": tr_capacity_kva, "expected_load_rate": round(expected_rate, 2), "is_safe": is_safe, "analysis": analysis}

    def step3_generate_boq(self, breaker_name: str, cable_sq: float, length_m: float, labor_unit_cost: int = 280000, custom_breaker_cost: int = 0, custom_cable_cost_per_m: int = 0, custom_tray_cost_per_m: int = 0) -> dict:
        breaker_cost = custom_breaker_cost if custom_breaker_cost > 0 else self.price_db.get(breaker_name, 120000)
        sq_val = cable_sq if isinstance(cable_sq, (int, float)) else 300
        
        cable_cost_per_m = custom_cable_cost_per_m if custom_cable_cost_per_m > 0 else (sq_val * 400)
        cable_cost = cable_cost_per_m * length_m
        
        tray_cost_per_m = custom_tray_cost_per_m if custom_tray_cost_per_m > 0 else self.price_db["Cable_Tray_W300"]
        tray_cost = tray_cost_per_m * length_m
        
        material_cost = breaker_cost + cable_cost + tray_cost

        man_days = (length_m / 10) * (1 + (sq_val / 100))
        labor_cost = int(man_days * labor_unit_cost)

        return {
            "applied_material_costs": {"breaker": breaker_cost, "cable_per_m": cable_cost_per_m, "tray_per_m": tray_cost_per_m},
            "material_cost": material_cost,
            "labor_cost": labor_cost,
            "total_estimated_cost": material_cost + labor_cost,
            "applied_labor_unit_cost": labor_unit_cost,
            "estimated_man_days": round(man_days, 1)
        }

engine = UtilityEngineeringEngine()

def precision_design_tool(load_kw: float, distance_m: float, power_factor: float = 0.9, demand_factor: float = 1.0, is_continuous: bool = True, power_type: str = "3상4선", install_method: str = "E", temp_c: float = 30.0, is_critical_load: bool = False, tr_capacity_kva: float = 1000.0) -> dict:
    """장비 용량, 거리 등과 함께 주요 공정/안전 설비 여부(is_critical_load)와 접속되는 변압기 용량(tr_capacity_kva)을 입력받아 UPS 연계 및 단락전류를 고려한 정밀 Sizing을 수행합니다."""
    return engine.step1_advanced_sizing(load_kw, distance_m, power_factor, demand_factor, is_continuous, power_type, install_method, temp_c, is_critical_load, tr_capacity_kva)

def evaluate_capacity_tool(tr_capacity_kva: float, current_load_kw: float, add_power_kw: float) -> dict:
    return engine.step2_evaluate_capacity(tr_capacity_kva, current_load_kw, add_power_kw)

def generate_boq_tool(breaker_name: str, cable_sq: float, length_m: float, labor_unit_cost: int = 280000, custom_breaker_cost: int = 0, custom_cable_cost_per_m: int = 0, custom_tray_cost_per_m: int = 0) -> dict:
    return engine.step3_generate_boq(breaker_name, cable_sq, length_m, labor_unit_cost, custom_breaker_cost, custom_cable_cost_per_m, custom_tray_cost_per_m)

tools_list = [precision_design_tool, evaluate_capacity_tool, generate_boq_tool]

# ==========================================
# 2. 문서 자동생성 로직
# ==========================================
def generate_excel_document(prompt_text, ai_response_text):
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    align_center = Alignment(horizontal="center", vertical="center")

    ws1 = wb.active
    ws1.title = "1. 증설 공사 기안 및 발주서"
    ws1.append(["항목", "내용"])
    for cell in ws1[1]: cell.font, cell.fill, cell.alignment = header_font, header_fill, align_center
    ws1.append(["요청 요약", prompt_text])
    ws1.append(["AI 최종 솔루션 요약", ai_response_text[:3000]])
    ws1.column_dimensions['A'].width = 25; ws1.column_dimensions['B'].width = 100

    ws2 = wb.create_sheet(title="2. 안전작업허가서(PTW)")
    ws2.append(["구분", "안전 확보 지침 (LOTO)"])
    for cell in ws2[1]: cell.font, cell.fill, cell.alignment = header_font, header_fill, align_center
    ws2.append(["작업 위험성 평가", "감전, 단락 사고 위험, 비상 설비 정전 위험"])
    ws2.append(["LOTO (차단 절차)", "1. 메인 판넬 차단기(MCCB) Open\n2. 잠금장치(Lock) 체결 및 위험 Tag 부착"])
    ws2.column_dimensions['A'].width = 25; ws2.column_dimensions['B'].width = 80

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def process_multimodal_file(file_obj, title_text):
    text_context, media_item = "", None
    if file_obj:
        if file_obj.name.endswith(('.csv', '.xlsx')):
            df = pd.read_csv(file_obj) if file_obj.name.endswith('.csv') else pd.read_excel(file_obj)
            text_context = f"\n\n### [{title_text}]\n{df.to_string(index=False)}"
        elif file_obj.type == "application/pdf":
            text_context = f"\n\n### [{title_text}]\n(첨부된 PDF 문서 시각적 데이터 참조)"
            media_item = {"mime_type": "application/pdf", "data": file_obj.getvalue()}
        else:
            text_context = f"\n\n### [{title_text}]\n(첨부된 이미지 시각적 데이터 참조)"
            media_item = Image.open(file_obj)
    return text_context, media_item

# ==========================================
# 3. Streamlit UI
# ==========================================
st.set_page_config(page_title="신입 엔지니어 필수 AI Agent", page_icon="⚡", layout="wide")
st.title("⚡ Utility 전기 AI Agent (KEC 정밀 + 고장전류/UPS 검증)")

with st.sidebar:
    st.header("🛠️ 업무 모드 선택")
    app_mode = st.radio("수행할 업무를 선택하세요:", ["🏗️ 증설 엔지니어링 (발주/안전)", "👨‍🏫 신입사원 SLD 튜터링"])
    st.markdown("---")
    
    st.header("📂 현장 데이터 업로드")
    sld_file = st.file_uploader("1. 단선도(SLD) 도면", type=["jpg", "png", "pdf"])
    if sld_file and sld_file.type != "application/pdf": st.image(Image.open(sld_file), use_container_width=True)

    scada_file = st.file_uploader("2. SCADA 부하 데이터 (선택)", type=["csv", "xlsx", "jpg", "png", "pdf"])
    scada_context, scada_media = process_multimodal_file(scada_file, "업로드된 SCADA 실시간 데이터 참고")
    if not scada_file: scada_context = "\n\n### [SCADA 데이터 부재]\n단선도(SLD)를 분석해 기존 부하를 100% 가동 조건으로 산정하세요."
    
    labor_file = st.file_uploader("3. 노무비 단가표 (선택)", type=["csv", "xlsx", "jpg", "png", "pdf"])
    labor_context, labor_media = process_multimodal_file(labor_file, "업로드된 노무비 단가표 참고")
    if not labor_file: labor_context = "\n\n### [노무비 단가 정보]\n업로드 없음. 2026년 표준 내선전공 단가(280,000원/인) 적용 요망."

    load_schedule_file = st.file_uploader("4. 증설 부하 리스트 (선택)", type=["csv", "xlsx", "jpg", "png", "pdf"])
    load_schedule_context, load_schedule_media = process_multimodal_file(load_schedule_file, "업로드된 증설 부하 리스트 참고")
    if not load_schedule_file: load_schedule_context = "\n\n### [증설 부하 리스트]\n업로드 없음. 사용자의 채팅 프롬프트에서 증설 계획을 파악하세요."

    material_file = st.file_uploader("5. 자재 단가표 (선택)", type=["csv", "xlsx", "jpg", "png", "pdf"])
    material_context, material_media = process_multimodal_file(material_file, "업로드된 자재 단가표 참고")
    if not material_file: material_context = "\n\n### [자재 단가 정보]\n업로드 없음. 시스템 내장 기본 자재 단가를 적용하세요."

if "eng_msg" not in st.session_state: st.session_state.eng_msg = []
if "tutor_msg" not in st.session_state: st.session_state.tutor_msg = []
if "latest_eng_result" not in st.session_state: st.session_state.latest_eng_result = None
if "latest_eng_prompt" not in st.session_state: st.session_state.latest_eng_prompt = None

if app_mode == "🏗️ 증설 엔지니어링 (발주/안전)":
    st.subheader("📊 부하 증설 검토 및 공사 발주 자동화")
    st.info("장비 증설 시 KEC 정밀 케이블 Sizing은 물론, **단락전류(kA) 검증 및 UPS(무정전전원) 필요 여부**까지 AI가 판단합니다.")
    
    for msg in st.session_state.eng_msg:
        with st.chat_message(msg["role"]): st.markdown(msg["content"])
        
    if prompt := st.chat_input("예: 가스 감지기 및 스크러버(중요 설비) 50kW 100m 증설. TR용량 1000kVA."):
        st.session_state.eng_msg.append({"role": "user", "content": prompt})
        with st.chat_message("user"): st.markdown(prompt)

        if API_KEY:
            with st.chat_message("assistant"):
                with st.spinner("계통 데이터 분석 및 정밀 규격 산출 중..."):
                    genai.configure(api_key=API_KEY)
                    sys_instruct = """당신은 반도체 Fab Hook-Up 공사를 총괄하는 최고 수준의 전기 엔지니어입니다.
                    다음 순서로 검토를 진행하세요:
                    1. evaluate_capacity_tool 호출 (SCADA 또는 도면 기반 부하 산정)
                    2. precision_design_tool 호출 시, 요청된 장비가 '가스감지기, 스크러버, 핵심공정장비' 등 중요/중단불가 설비로 판단되면 is_critical_load=True 로 설정하여 UPS 연계 여부를 확인하고, 해당 변압기 용량을 tr_capacity_kva 에 입력하여 고장전류(kA)를 반영하세요.
                    3. generate_boq_tool 호출 시 [노무비/자재 단가 정보] 참조
                    
                    최종 보고서 목차:
                    
                    ### 1. 변압기 부하 분석 및 UPS 연계 검토
                    - SCADA/SLD 기반 부하율 분석
                    - ⚡ [신규 기능] 증설 장비 중요도(Critical 여부) 평가 및 UPS/C-UPS 연계 필요성
                    
                    ### 2. KEC 정밀 Sizing 및 고장전류 검증
                    - 최적 케이블 SQ 및 전압강하율
                    - ⚡ [신규 기능] 변압기 용량 기준 단락전류 계산 및 차단기 필요 차단용량(kA) 명시
                    
                    ### 3. 🚨 주요 문제점 및 맞춤형 대안 (VE)
                    - [문제 1: 내용] 👉 [대안 1: 내용]
                    - [문제 2: 내용] 👉 [대안 2: 내용]
                    (문제가 없다면 '특이사항 없음' 기재)
                    
                    ### 4. 발주 예상 공사비
                    - 자재비 및 노무비(단가 출처 명시)
                    
                    ### 5. 🎯 최종 합리적 증설 솔루션 요약
                    - 계통, 케이블, 차단용량(kA), UPS 여부를 종합한 한 문단 결론"""
                    
                    model = genai.GenerativeModel(model_name='gemini-3.6-flash', system_instruction=sys_instruct, tools=tools_list)
                    chat = model.start_chat(enable_automatic_function_calling=True)
                    
                    contents = []
                    if sld_file:
                        if sld_file.type == "application/pdf": contents.append({"mime_type": "application/pdf", "data": sld_file.getvalue()})
                        else: contents.append(Image.open(sld_file))
                    
                    for media_item in [scada_media, labor_media, load_schedule_media, material_media]:
                        if media_item: contents.append(media_item)

                    contents.append(prompt + scada_context + labor_context + load_schedule_context + material_context)

                    retry_delay = 5
                    for attempt in range(3):
                        try:
                            res = chat.send_message(contents)
                            final_text = res.text
                            break
                        except Exception as e:
                            if "503" in str(e) or "429" in str(e):
                                if attempt < 2: time.sleep(retry_delay); retry_delay *= 2
                                else: raise Exception("서버 혼잡.")
                            else: raise e
                    
                    st.markdown(final_text)
                    st.session_state.eng_msg.append({"role": "assistant", "content": final_text})
                    st.session_state.latest_eng_result = final_text
                    st.session_state.latest_eng_prompt = prompt

    if st.session_state.latest_eng_result:
        st.markdown("---")
        st.subheader("📑 결재용 문서 미리보기 및 다운로드")
        
        df_draft = pd.DataFrame([
            ["공사명", "Utility 설비 Hook-up 증설 공사"],
            ["요청 요약", st.session_state.latest_eng_prompt],
            ["최종 솔루션", "위의 [5. 최종 합리적 증설 솔루션 요약] 참조"]
        ], columns=["항목", "내용"])
        
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("#### 📄 1. 증설 기안서 요약")
            st.table(df_draft)
        with col2:
            st.markdown("#### 📄 2. 안전작업허가서(PTW)")
            st.table(pd.DataFrame([["위험성 평가", "감전, 단락 사고 위험, 비상 정전 위험"], ["LOTO 절차", "1. 차단기 Open 2. 잠금 및 Tag 부착"]], columns=["구분", "지침"]))

        excel_data = generate_excel_document(st.session_state.latest_eng_prompt, st.session_state.latest_eng_result)
        st.download_button("📥 통합 보고서 다운로드 (엑셀)", data=excel_data, file_name="HookUp_최종결정서.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

else:
    st.subheader("👨‍🏫 SLD 도면 스터디 (친절한 사수 AI)")
    for msg in st.session_state.tutor_msg:
        with st.chat_message(msg["role"]): st.markdown(msg["content"])
        
    if prompt := st.chat_input("도면에 대해 질문하세요"):
        st.session_state.tutor_msg.append({"role": "user", "content": prompt})
        with st.chat_message("user"): st.markdown(prompt)

        if API_KEY:
            with st.chat_message("assistant"):
                with st.spinner("사수 AI가 도면을 확인 중..."):
                    genai.configure(api_key=API_KEY)
                    sys_instruct = "당신은 신입사원에게 도면(SLD)을 가르쳐주는 10년 차 사수입니다."
                    model = genai.GenerativeModel(model_name='gemini-3.6-flash', system_instruction=sys_instruct)
                    
                    contents = []
                    if sld_file:
                        if sld_file.type == "application/pdf": contents.append({"mime_type": "application/pdf", "data": sld_file.getvalue()})
                        else: contents.append(Image.open(sld_file))
                    contents.append(prompt)

                    retry_delay = 5
                    for attempt in range(3):
                        try:
                            final_text = model.generate_content(contents).text
                            break
                        except Exception as e:
                            if "503" in str(e) or "429" in str(e):
                                if attempt < 2: time.sleep(retry_delay); retry_delay *= 2
                                else: raise Exception("서버 혼잡.")
                            else: raise e
                    
                    st.markdown(final_text)
                    st.session_state.tutor_msg.append({"role": "assistant", "content": final_text})